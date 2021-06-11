from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import os
import sqlite3

def create_excel(face_dict,subject):
    #database connection
    conn = sqlite3.connect('./Face-Database')
    c = conn.cursor()

    #get current date
    Date = date.today()
    currentDate = Date.strftime("%d_%m_%y")
    Time = datetime.now()
    currentTime = Time.strftime("%H:%M:%S")
    currentMonth = Time.strftime("%B")
    currentYear = Time.strftime("%Y")
    currentDate = currentDate.replace('_','-')
    c.execute("SELECT * FROM Students ORDER BY Roll ASC")
    Filename = "./Attendance/"+subject+"/Attendance "+currentMonth+" "+currentYear+".xlsx"
    #create a workbook and add a worksheet
    if(os.path.exists(Filename)):
        wb = load_workbook(filename = Filename)
        ws1 = wb.active
        #for row_cells in ws1.iter_rows(min_row=1,max_row=1):
         #   print(row_cells)
        i=6
        while i>0:
            if ws1.cell(row=1,column=i).value is None:
                break
            i+=1
        col_name=ws1.cell(row=1,column=i).column_letter
        #print(col_name)
        #if currentDate==ws1.cell(row=1,column=i-1).value:
        #    col_name=ws1.cell(row=1,column=i-1).column_letter
            #print(col_name)
        ws1[col_name+"1"] = currentDate
        ws1[col_name+"2"] = currentTime
        for i in range(3,ws1.max_row+1):
            if ws1["B"+str(i)].value in face_dict:
                ws1[col_name+str(i)] = 'P'
            else:
                ws1[col_name+str(i)] = 'A'
        Stud_att = {}
        for i in range(3,ws1.max_row+1):
            for j in range(6, ws1.max_column+1):
                col=ws1.cell(row=i,column=j).column_letter
                if ws1[col + str(i)].value == "P":
                    Stud_att[ws1["B" + str(i)].value] = Stud_att.get(ws1["B" + str(i)].value, 0) + 1
                else:
                    if ws1["B" + str(i)].value not in Stud_att:
                        Stud_att[ws1["B" + str(i)].value] = 0
        l=3
        for (k, v) in Stud_att.items():
            Stud_att[k] = (Stud_att[k] / (ws1.max_column-5)) * 100
            ws1["D"+str(l)] = Stud_att[ws1["B"+str(l)].value]
            l+=1
        wb.save(filename=Filename)
    else:
        if not os.path.exists("./Attendance/"+subject):
            os.makedirs("./Attendance/"+subject)
        wb = Workbook()
        dest_filename = "./Attendance/"+subject+"/Attendance "+currentMonth+" "+currentYear+".xlsx"

        #creating worksheet and giving names to column
        ws1 = wb.active
        ws1.title = subject
        ws1.append(('Roll Number', 'Name', '', '% Attendance',''))
        ws1.append(('', '', '', '',''))

    #entering students information from database
        #for row_cells in ws1.iter_rows(min_row=1,max_row=1):
         #   print(row_cells)
        ws1["F1"]=currentDate
        ws1["F2"]=currentTime
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                if a[1] in face_dict:
                    ws1.append((a[2], a[1],'',100,'','P'))
                else:
                    ws1.append((a[2],a[1],'',0,'','A'))
        #saving the file
        wb.save(filename=dest_filename)
