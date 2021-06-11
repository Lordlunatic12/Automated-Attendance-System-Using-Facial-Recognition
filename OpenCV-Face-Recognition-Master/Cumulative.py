from openpyxl import Workbook, load_workbook
import sqlite3
import os
import time

def listdefaulter(subject):
    folderName = subject
    folderPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Attendance/"+folderName)
    '''if not os.path.exists(folderPath):
        os.makedirs(folderPath)'''
    currentDate = time.strftime("%d_%m_%y")
    currentDate = currentDate.replace('_', '-')
    def_path = folderPath + "/Defaulter's Attendance"
    if not os.path.exists(def_path):
        os.makedirs(def_path)
    dest_filename = def_path + "/Defaulter's List " + currentDate + ".xlsx"
    if os.path.exists(dest_filename):
        return -1
    else:
        Stud_att = {}
        conn = sqlite3.connect('./Face-Database')
        c = conn.cursor()
        c.execute("SELECT * FROM Students ORDER BY Roll ASC")

        path = './Attendance/'+subject
        AttFiles = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.xlsx')]
        #print(AttFiles)
        for Attfile in AttFiles:
            wb = load_workbook(filename=Attfile)
            sheet = wb.active
            for i in range(3,sheet.max_row+1):
                Stud_att[sheet["B"+str(i)].value] = Stud_att.get(sheet["B"+str(i)].value,0)+sheet["D"+str(i)].value
        for (k,v) in Stud_att.items():
            Stud_att[k] = (Stud_att[k]/len(AttFiles))
        #print(Stud_att)
        wb = Workbook()
        #creating worksheet and giving names to column
        ws1 = wb.active
        ws1.title = subject
        ws1.append(('Roll Number', 'Defaulter Name', '% Attendance'))
        ws1.append(('', '', '', ''))

        #entering students information from database
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                if a[1] in Stud_att:
                    if Stud_att[a[1]]<=75:
                        ws1.append((a[2], a[1],Stud_att[a[1]]))
            #saving the file
        wb.save(filename = dest_filename)
        return 0

def markscalc(subject):
    folderName = subject
    folderPath = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Attendance/" + folderName)
    '''if not os.path.exists(folderPath):
        os.makedirs(folderPath)'''
    currentDate = time.strftime("%d_%m_%y")
    currentDate = currentDate.replace('_', '-')
    def_path = folderPath + "/Attendance Marks"
    if not os.path.exists(def_path):
        os.makedirs(def_path)
    dest_filename = def_path + "/Attendance Marks" + currentDate + ".xlsx"
    if os.path.exists(dest_filename):
        return -1
    else:
        Stud_att = {}
        conn = sqlite3.connect('./Face-Database')
        c = conn.cursor()
        c.execute("SELECT * FROM Students ORDER BY Roll ASC")

        path = './Attendance/' + subject
        AttFiles = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.xlsx')]
        # print(AttFiles)
        for Attfile in AttFiles:
            wb = load_workbook(filename=Attfile)
            sheet = wb.active
            for i in range(3, sheet.max_row + 1):
                Stud_att[sheet["B" + str(i)].value] = Stud_att.get(sheet["B" + str(i)].value, 0) + sheet[
                    "D" + str(i)].value
        for (k, v) in Stud_att.items():
            Stud_att[k] = (Stud_att[k] / len(AttFiles))
        # print(Stud_att)
        wb = Workbook()
        # creating worksheet and giving names to column
        ws1 = wb.active
        ws1.title = subject
        ws1.append(('Roll Number', 'Name', '% Attendance','Marks'))
        ws1.append(('', '', '', ''))

        # entering students information from database
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                if a[1] in Stud_att:
                    if Stud_att[a[1]] >= 95:
                        ws1.append((a[2], a[1], Stud_att[a[1]],5))
                    elif Stud_att[a[1]] >= 90:
                        ws1.append((a[2], a[1], Stud_att[a[1]],4))
                    elif Stud_att[a[1]] >= 85:
                        ws1.append((a[2], a[1], Stud_att[a[1]],3))
                    elif Stud_att[a[1]] >= 80:
                        ws1.append((a[2], a[1], Stud_att[a[1]],2))
                    elif Stud_att[a[1]] >= 75:
                        ws1.append((a[2], a[1], Stud_att[a[1]],1))
                    else:
                        ws1.append((a[2], a[1], Stud_att[a[1]],0))
            # saving the file
        wb.save(filename=dest_filename)
        return 0
